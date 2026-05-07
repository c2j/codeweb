#!/usr/bin/env python3
"""Convert AAS lineage Excel file to CGEF JSON format."""

import argparse
import hashlib
import json
import re
import sys
import warnings
from datetime import datetime, timezone
from pathlib import Path


def slugify(s: str) -> str:
    if not s:
        return ""
    for ext in (".op", ".java"):
        if s.lower().endswith(ext):
            s = s[: -len(ext)]
            break
    result = re.sub(r"[^a-zA-Z0-9]+", "_", s)
    result = result.strip("_")
    return result.lower()


def hash_id(s: str) -> str:
    """Stable hash for use in node IDs (deterministic across runs)."""
    return hashlib.md5(s.encode()).hexdigest()[:12]


def parse_pkgproc_name(pkgproc_name: str):
    if not pkgproc_name:
        return None, None
    pkgproc_name = pkgproc_name.strip()
    if not pkgproc_name:
        return None, None
    if "." in pkgproc_name:
        package, name = pkgproc_name.split(".", 1)
        return package.lower(), name.lower()
    return None, pkgproc_name.lower()


def extract_namespace(mybatis_id: str) -> str:
    if not mybatis_id:
        return ""
    parts = mybatis_id.strip().split(".")
    if len(parts) >= 2:
        return parts[0]
    return mybatis_id.strip()


def extract_statement_id(mybatis_id: str) -> str:
    if not mybatis_id:
        return ""
    parts = mybatis_id.strip().split(".")
    if len(parts) >= 2:
        return ".".join(parts[1:])
    return mybatis_id.strip()


class CGEFBuilder:
    def __init__(self):
        self.nodes = {}
        self.edges = set()
        self.node_list = []
        self.edge_list = []
        self.warnings = []

    def add_node(self, node_id: str, node_type: str, key: dict, properties: dict = None, location: dict = None):
        if node_id in self.nodes:
            return False
        node = {
            "id": node_id,
            "type": node_type,
            "key": key,
        }
        if properties:
            node["properties"] = properties
        if location:
            node["location"] = location
        self.nodes[node_id] = node
        self.node_list.append(node)
        return True

    def add_edge(self, source: str, target: str, edge_type: str, properties: dict = None, location: dict = None):
        key = (source, target, edge_type)
        if key in self.edges:
            return False
        self.edges.add(key)
        edge = {
            "source": source,
            "target": target,
            "type": edge_type,
        }
        if properties:
            edge["properties"] = properties
        if location:
            edge["location"] = location
        self.edge_list.append(edge)
        return True

    def warn(self, msg: str):
        self.warnings.append(msg)
        warnings.warn(msg)

    def build(self) -> dict:
        return {
            "format_version": 1,
            "metadata": {
                "source": "aas-lineage-excel",
                "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                "description": "AAS系统血缘关系数据",
            },
            "node_schemas": {
                "menu": {
                    "display_name": "菜单",
                    "key_fields": ["id"],
                    "properties": {
                        "menu_path": {"type": "string", "description": "菜单面包屑路径"},
                        "resource_value_name": {"type": "string", "description": "菜单显示名称"},
                        "parent_value": {"type": "string", "description": "父菜单ID"},
                        "resource_id": {"type": "string", "description": "资源标识"},
                    },
                },
                "op_handler": {
                    "display_name": "OP处理器",
                    "key_fields": ["op_name"],
                    "properties": {
                        "servlet_url": {"type": "string", "description": "Servlet调用地址"},
                    },
                },
                "jsp": {
                    "display_name": "JSP页面",
                    "key_fields": ["jsp_name"],
                    "properties": {
                        "path": {"type": "string", "description": "JSP文件路径"},
                    },
                },
                "ctp_op_config": {
                    "display_name": "CTP OP配置",
                    "key_fields": ["op_file_name"],
                    "properties": {
                        "step_type": {"type": "string", "description": "步骤类型"},
                        "src_value": {"type": "string", "description": "OP服务引用"},
                    },
                },
                "java_op_step": {
                    "display_name": "Java OpStep",
                    "key_fields": ["class_name"],
                    "properties": {},
                },
            },
            "edge_schemas": {
                "menu_triggers_op": {
                    "display_name": "菜单触发操作",
                    "source_types": ["menu"],
                    "target_types": ["op_handler"],
                },
                "menu_parent": {
                    "display_name": "菜单层级",
                    "source_types": ["menu"],
                    "target_types": ["menu"],
                },
                "op_calls_procedure": {
                    "display_name": "OP调用存储过程",
                    "source_types": ["op_handler"],
                    "target_types": ["procedure", "unresolved"],
                },
                "op_renders_jsp": {
                    "display_name": "OP渲染JSP",
                    "source_types": ["op_handler"],
                    "target_types": ["jsp"],
                },
                "ctp_step_calls_procedure": {
                    "display_name": "CTP步骤调用存储过程",
                    "source_types": ["ctp_op_config"],
                    "target_types": ["procedure", "unresolved"],
                },
                "java_step_calls_procedure": {
                    "display_name": "Java类调用存储过程",
                    "source_types": ["java_op_step"],
                    "target_types": ["procedure", "unresolved"],
                },
            },
            "nodes": self.node_list,
            "edges": self.edge_list,
        }


def col_map(headers):
    return {name: i for i, name in enumerate(headers)}


def cell_value(row, col_map, name, default=""):
    idx = col_map.get(name)
    if idx is None:
        return default
    val = row[idx]
    return str(val) if val is not None else default


def process_sheet1(builder: CGEFBuilder, ws):
    headers = [cell.value for cell in ws[1]]
    cm = col_map(headers)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        resource_value = cell_value(row, cm, "RESOURCE_VALUE")
        if not resource_value:
            builder.warn(f"Sheet1 row {row_idx}: missing RESOURCE_VALUE, skipping")
            continue

        rv_str = resource_value.lstrip("-")
        menu_id = f"menu_{rv_str}"
        menu_path = cell_value(row, cm, "FNC_GET_MENU_PATH")
        resource_value_name = cell_value(row, cm, "RESOURCE_VALUE_NAME")
        parent_value = cell_value(row, cm, "PARENT_VALUE")
        resource_id = cell_value(row, cm, "RESOURCE_ID")
        op_name = cell_value(row, cm, "OP_NAME")
        jsp_name = cell_value(row, cm, "JSP_NAME")
        jsp_path = cell_value(row, cm, "JSP_PATH")
        pkgproc_name = cell_value(row, cm, "PKGPROC_NAME")
        comm = cell_value(row, cm, "COMM")

        builder.add_node(
            menu_id,
            "menu",
            key={
                "id": resource_value,
                "menu_path": menu_path,
                "name": resource_value_name,
            },
            properties={
                "menu_path": menu_path,
                "resource_value_name": resource_value_name,
                "parent_value": parent_value,
                "resource_id": resource_id,
            },
        )

        if op_name:
            op_id = f"op_{op_name}"
            builder.add_node(
                op_id,
                "op_handler",
                key={"op_name": op_name},
                properties={"servlet_url": comm},
            )
            builder.add_edge(menu_id, op_id, "menu_triggers_op")

            if jsp_path:
                jsp_id = f"jsp_{hash_id(jsp_path)}"
                builder.add_node(
                    jsp_id,
                    "jsp",
                    key={"jsp_name": jsp_name, "path": jsp_path},
                    properties={"path": jsp_path},
                )
                builder.add_edge(op_id, jsp_id, "op_renders_jsp")

            if pkgproc_name:
                pkg, name = parse_pkgproc_name(pkgproc_name)
                if pkg and name:
                    sp_id = f"sp_{pkg}_{name}"
                    builder.add_node(
                        sp_id,
                        "procedure",
                        key={"package": pkg, "name": name},
                        location={"file": "lineage/sheet1-menu-op", "line": 1},
                    )
                elif name:
                    sp_id = f"unres_{name}"
                    builder.add_node(
                        sp_id,
                        "unresolved",
                        key={"raw_expr": pkgproc_name, "context": op_name},
                    )
                else:
                    builder.warn(f"Sheet1 row {row_idx}: empty pkgproc_name after parse")
                    continue
                builder.add_edge(op_id, sp_id, "op_calls_procedure")

        if parent_value:
            parent_rv = parent_value.lstrip("-")
            parent_id = f"menu_{parent_rv}"
            if parent_id in builder.nodes:
                builder.add_edge(menu_id, parent_id, "menu_parent")


def process_sheet2(builder: CGEFBuilder, ws):
    headers = [cell.value for cell in ws[1]]
    cm = col_map(headers)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        op_name = cell_value(row, cm, "OP_NAME")
        if not op_name:
            builder.warn(f"Sheet2 row {row_idx}: missing OP_NAME, skipping")
            continue

        java_id = f"java_{slugify(op_name)}"
        builder.add_node(java_id, "java_op_step", key={"class_name": op_name})

        pkgproc_name = cell_value(row, cm, "PKGPROC_NAME")
        if pkgproc_name:
            pkg, name = parse_pkgproc_name(pkgproc_name)
            if pkg and name:
                sp_id = f"sp_{pkg}_{name}"
                builder.add_node(
                    sp_id,
                    "procedure",
                    key={"package": pkg, "name": name},
                    location={"file": "lineage/sheet2-java-op", "line": 1},
                )
            elif name:
                sp_id = f"unres_{name}"
                builder.add_node(
                    sp_id,
                    "unresolved",
                    key={"raw_expr": pkgproc_name, "context": op_name},
                )
            else:
                builder.warn(f"Sheet2 row {row_idx}: empty pkgproc_name after parse")
                continue
            builder.add_edge(java_id, sp_id, "java_step_calls_procedure")

            mybatis_id = cell_value(row, cm, "MYBATIS_ID")
            if mybatis_id and mybatis_id != "—":
                ms_id = f"ms_{slugify(mybatis_id)}"
                ns = extract_namespace(mybatis_id)
                stmt_id = extract_statement_id(mybatis_id)
                mybatis_file_path = cell_value(row, cm, "MYBATIS_FILE_PATH")
                builder.add_node(
                    ms_id,
                    "mapped_statement",
                    key={"namespace": ns, "statement_id": stmt_id, "kind": "select"},
                    location={
                        "file": mybatis_file_path if mybatis_file_path else "lineage/sheet2-java-op",
                        "line": 1,
                    },
                )
                builder.add_edge(java_id, ms_id, "invokes_mapper")
                builder.add_edge(ms_id, sp_id, "calls_procedure")


def process_sheet3(builder: CGEFBuilder, ws):
    headers = [cell.value for cell in ws[1]]
    cm = col_map(headers)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        op_file_name = cell_value(row, cm, "OP_FILE_NAME")
        if not op_file_name:
            builder.warn(f"Sheet3 row {row_idx}: missing OP_FILE_NAME, skipping")
            continue

        ctp_id = f"ctp_{slugify(op_file_name)}"
        type_value = cell_value(row, cm, "TYPE_VALUE")
        src_value = cell_value(row, cm, "SRC_VALUE")
        builder.add_node(
            ctp_id,
            "ctp_op_config",
            key={"op_file_name": op_file_name},
            properties={"step_type": type_value, "src_value": src_value},
        )

        pkgproc_name = cell_value(row, cm, "PKGPROC_NAME")
        if pkgproc_name:
            pkg, name = parse_pkgproc_name(pkgproc_name)
            if pkg and name:
                sp_id = f"sp_{pkg}_{name}"
                builder.add_node(
                    sp_id,
                    "procedure",
                    key={"package": pkg, "name": name},
                    location={"file": "lineage/sheet3-ctp-op", "line": 1},
                )
            elif name:
                sp_id = f"unres_{name}"
                builder.add_node(
                    sp_id,
                    "unresolved",
                    key={"raw_expr": pkgproc_name, "context": op_file_name},
                )
            else:
                builder.warn(f"Sheet3 row {row_idx}: empty pkgproc_name after parse")
                continue
            builder.add_edge(ctp_id, sp_id, "ctp_step_calls_procedure")


def process_sheet4(builder: CGEFBuilder, ws):
    headers = [cell.value for cell in ws[1]]
    cm = col_map(headers)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        resource_value = cell_value(row, cm, "RESOURCE_VALUE")
        if not resource_value:
            builder.warn(f"Sheet4 row {row_idx}: missing RESOURCE_VALUE, skipping")
            continue

        rv_str = resource_value.lstrip("-")
        menu_id = f"menu_{rv_str}"
        menu_path = cell_value(row, cm, "FNC_GET_MENU_PATH")
        resource_value_name = cell_value(row, cm, "RESOURCE_VALUE_NAME")
        parent_value = cell_value(row, cm, "PARENT_VALUE")
        resource_id = cell_value(row, cm, "RESOURCE_ID")
        op_name = cell_value(row, cm, "OP_NAME")
        comm = cell_value(row, cm, "COMM")

        builder.add_node(
            menu_id,
            "menu",
            key={
                "id": resource_value,
                "menu_path": menu_path,
                "name": resource_value_name,
            },
            properties={
                "menu_path": menu_path,
                "resource_value_name": resource_value_name,
                "parent_value": parent_value,
                "resource_id": resource_id,
            },
        )

        if op_name:
            op_id = f"op_{op_name}"
            builder.add_node(
                op_id,
                "op_handler",
                key={"op_name": op_name},
                properties={"servlet_url": comm},
            )
            builder.add_edge(menu_id, op_id, "menu_triggers_op")

        if parent_value:
            parent_rv = parent_value.lstrip("-")
            parent_id = f"menu_{parent_rv}"
            if parent_id in builder.nodes:
                builder.add_edge(menu_id, parent_id, "menu_parent")


def main():
    parser = argparse.ArgumentParser(description="Convert AAS lineage Excel to CGEF JSON")
    parser.add_argument("excel_path", help="Path to lineage Excel file")
    parser.add_argument("-o", "--output", required=True, help="Output CGEF JSON path")
    parser.add_argument("--validate", action="store_true", help="Validate output against CGEF schema")
    args = parser.parse_args()

    excel_path = Path(args.excel_path)
    output_path = Path(args.output)

    if not excel_path.exists():
        print(f"Error: Excel file not found: {excel_path}", file=sys.stderr)
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl is required but not installed", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    expected_sheets = ['菜单OP分析结果', 'com.icbc.aas OP分析结果', 'com.icbc.ctp OP分析结果', 'op菜单配置清单表']
    if sheet_names != expected_sheets:
        print(f"Warning: Unexpected sheet order or names: {sheet_names}", file=sys.stderr)

    builder = CGEFBuilder()

    process_sheet1(builder, wb['菜单OP分析结果'])
    process_sheet2(builder, wb['com.icbc.aas OP分析结果'])
    process_sheet3(builder, wb['com.icbc.ctp OP分析结果'])
    process_sheet4(builder, wb['op菜单配置清单表'])

    cgef = builder.build()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(cgef, f, ensure_ascii=False, indent=2)

    print(f"Written {output_path}")
    print(f"  Nodes: {len(cgef['nodes'])}")
    print(f"  Edges: {len(cgef['edges'])}")

    node_counts = {}
    for node in cgef["nodes"]:
        t = node["type"]
        node_counts[t] = node_counts.get(t, 0) + 1
    edge_counts = {}
    for edge in cgef["edges"]:
        t = edge["type"]
        edge_counts[t] = edge_counts.get(t, 0) + 1

    print("\nNode counts by type:")
    for t, count in sorted(node_counts.items()):
        print(f"  {t}: {count}")
    print("\nEdge counts by type:")
    for t, count in sorted(edge_counts.items()):
        print(f"  {t}: {count}")

    if builder.warnings:
        print(f"\nWarnings ({len(builder.warnings)}):")
        for w in builder.warnings[:10]:
            print(f"  {w}")
        if len(builder.warnings) > 10:
            print(f"  ... and {len(builder.warnings) - 10} more")

    if args.validate:
        try:
            import jsonschema
        except ImportError:
            print("\nWarning: jsonschema not installed, skipping validation", file=sys.stderr)
            return

        schema_path = Path(__file__).parent.parent / "docs" / "cgef-schema.json"
        if not schema_path.exists():
            schema_path = Path("docs/cgef-schema.json")
        if not schema_path.exists():
            print(f"\nWarning: Schema not found at {schema_path}, skipping validation", file=sys.stderr)
            return

        with open(schema_path, "r", encoding="utf-8") as f:
            schema = json.load(f)

        try:
            jsonschema.validate(cgef, schema)
            print("\nValidation: PASSED")
        except jsonschema.ValidationError as e:
            print(f"\nValidation FAILED: {e.message}", file=sys.stderr)
            print(f"  at: {'/'.join(str(p) for p in e.path)}", file=sys.stderr)
            sys.exit(1)


if __name__ == "__main__":
    main()
